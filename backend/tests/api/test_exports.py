"""Exporting a run.

The rule every case here checks: **an export carries the manifest that
reproduces it, or it ships beside one.** An exported number that cannot be
traced back to the data and the tool version that produced it is exactly what
this project exists not to produce, and a file outlives the app that made it —
once it is on someone's disk, the provenance has to travel with it.
"""

import csv
import io
import json
import zipfile
from datetime import date
from uuid import uuid4

import numpy as np
import pandas as pd
import pytest
import pytest_asyncio
from openpyxl import load_workbook

from econometrica.api.deps import get_price_source, get_provider_registry
from econometrica.llm.fake import FakeProvider
from econometrica.main import app

QUESTION = "Does AAA follow a random walk?"

PLAN = {
    "question": QUESTION,
    "dataset": {"tickers": ["AAA"], "start": "2020-01-01", "end": "2020-06-30"},
    "steps": [
        {"id": "s1", "tool": "realized_vol", "params": {"column": "AAA_return", "window": 20}}
    ],
}
NARRATIVE = json.dumps({"prose": "The series wanders.", "citations": ["s1"]})


class FakeSource:
    async def prices(self, ticker: str, *, start: date, end: date) -> pd.Series:
        index = pd.date_range("2020-01-01", periods=182, freq="D")
        rng = np.random.default_rng(7)
        return pd.Series(100.0 + np.cumsum(rng.normal(size=182)), index=index)


class ScriptedRegistry:
    def __init__(self) -> None:
        self.provider = FakeProvider(name="ollama", responses=[json.dumps(PLAN), NARRATIVE])

    def spec(self, name: str) -> object:
        if name != "ollama":
            raise KeyError(name)
        return object()

    def is_configured(self, name: str) -> bool:
        return name == "ollama"

    def build(self, name: str) -> FakeProvider:
        return self.provider


@pytest_asyncio.fixture
async def scripted():
    registry = ScriptedRegistry()
    app.dependency_overrides[get_provider_registry] = lambda: registry
    app.dependency_overrides[get_price_source] = lambda: FakeSource()
    yield registry
    app.dependency_overrides.pop(get_provider_registry, None)
    app.dependency_overrides.pop(get_price_source, None)


async def finished_run(client) -> str:
    project = (await client.post("/api/projects", json={"name": "Exports"})).json()
    await client.patch(
        f"/api/projects/{project['id']}",
        json={
            "validation_tier": "single",
            "model_assignments": {
                role: {"provider": "ollama", "model": "fake-1"}
                for role in ("planner", "narrator")
            },
        },
    )
    chat = (
        await client.post(f"/api/projects/{project['id']}/chats", json={"name": "c"})
    ).json()
    await client.post(f"/api/chats/{chat['id']}/runs", json={"question": QUESTION})
    return str((await client.get(f"/api/chats/{chat['id']}/runs")).json()[0]["id"])


# --- the shape of the offer --------------------------------------------------


async def test_an_unknown_run_cannot_be_exported(client, scripted):
    response = await client.get(f"/api/runs/{uuid4()}/export", params={"format": "json"})
    assert response.status_code == 404


async def test_an_unknown_format_is_refused_by_name(client, scripted):
    run_id = await finished_run(client)

    response = await client.get(f"/api/runs/{run_id}/export", params={"format": "docx"})

    assert response.status_code == 422


@pytest.mark.parametrize(
    ("fmt", "content_type", "suffix"),
    [
        ("json", "application/json", ".json"),
        ("markdown", "text/markdown", ".md"),
        ("csv", "text/csv", ".csv"),
        ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
        ("zip", "application/zip", ".zip"),
    ],
)
async def test_every_format_downloads_as_a_named_file(client, scripted, fmt, content_type, suffix):
    run_id = await finished_run(client)

    response = await client.get(f"/api/runs/{run_id}/export", params={"format": fmt})

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(content_type)
    # Without a filename the browser saves it as the route's last segment.
    assert suffix in response.headers["content-disposition"]
    assert response.content


# --- the manifest travels with the file --------------------------------------


async def test_json_carries_the_whole_outcome_and_its_manifests(client, scripted):
    run_id = await finished_run(client)

    payload = (await client.get(f"/api/runs/{run_id}/export", params={"format": "json"})).json()

    assert payload["run"]["question"] == QUESTION
    result = payload["outcome"]["execution"]["outcomes"][0]["result"]
    assert result["manifest"]["tool"] == "realized_vol"
    assert result["manifest"]["data_fingerprint"]


async def test_markdown_reports_the_manifest_beside_the_findings(client, scripted):
    run_id = await finished_run(client)

    body = (await client.get(f"/api/runs/{run_id}/export", params={"format": "markdown"})).text

    assert QUESTION in body
    assert "realized_vol" in body
    # The reader of a pasted report must be able to ask "from what data?".
    assert "fingerprint" in body.lower()


async def test_csv_prepends_the_manifest_as_comments(client, scripted):
    # CSV has no metadata channel, so the provenance rides in comment lines —
    # which pandas skips with `comment="#"` and a person can simply read.
    run_id = await finished_run(client)

    body = (await client.get(f"/api/runs/{run_id}/export", params={"format": "csv"})).text

    assert body.startswith("#")
    assert "fingerprint" in body.lower()

    rows = list(csv.reader(line for line in body.splitlines() if not line.startswith("#")))
    assert rows[0] == ["step_id", "tool", "series", "x", "y"]
    assert len(rows) > 1


async def test_xlsx_keeps_a_manifest_sheet(client, scripted):
    run_id = await finished_run(client)

    response = await client.get(f"/api/runs/{run_id}/export", params={"format": "xlsx"})

    book = load_workbook(io.BytesIO(response.content))
    assert "manifest" in book.sheetnames
    assert any(sheet != "manifest" for sheet in book.sheetnames), "the data itself is missing"
    values = [str(cell.value) for row in book["manifest"].iter_rows() for cell in row]
    assert any("realized_vol" in value for value in values)


async def test_the_zip_carries_every_artifact_and_the_manifest(client, scripted):
    run_id = await finished_run(client)

    response = await client.get(f"/api/runs/{run_id}/export", params={"format": "zip"})

    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        names = set(archive.namelist())
        assert {"run.json", "report.md", "series.csv", "results.xlsx", "manifest.json"} <= names
        manifest = json.loads(archive.read("manifest.json"))

    # The manifest names what produced the run, not merely that one existed.
    assert manifest["run_id"] == run_id
    assert manifest["tools"][0]["tool"] == "realized_vol"
    assert manifest["tools"][0]["data_fingerprint"]


async def test_a_run_that_produced_nothing_still_exports(client, scripted):
    # A failed run is worth keeping: the export is how someone reports it.
    scripted.provider.responses = ["not json", "still not json"]
    project = (await client.post("/api/projects", json={"name": "Empty"})).json()
    await client.patch(
        f"/api/projects/{project['id']}",
        json={
            "validation_tier": "single",
            "model_assignments": {
                role: {"provider": "ollama", "model": "fake-1"}
                for role in ("planner", "narrator")
            },
        },
    )
    chat = (
        await client.post(f"/api/projects/{project['id']}/chats", json={"name": "c"})
    ).json()
    await client.post(f"/api/chats/{chat['id']}/runs", json={"question": QUESTION})
    run_id = (await client.get(f"/api/chats/{chat['id']}/runs")).json()[0]["id"]

    for fmt in ("json", "markdown", "csv", "xlsx", "zip"):
        response = await client.get(f"/api/runs/{run_id}/export", params={"format": fmt})
        assert response.status_code == 200, fmt
        assert response.content, fmt


async def test_a_refused_step_reports_no_result_rather_than_a_blank(client, scripted):
    # A refused step has no manifest because it produced nothing. Empty cells
    # would read as provenance that went missing.
    run_id = await finished_run(client)
    body = (await client.get(f"/api/runs/{run_id}/export", params={"format": "markdown"})).text

    for line in body.splitlines():
        if line.startswith("| s") and "no result" not in line:
            assert "``" not in line, f"blank provenance in: {line}"
