"""Turning a stored run into files someone can keep.

**The rule for every format here: the manifest travels with the export.** A
number on someone's disk outlives the app that produced it, and one that
cannot be traced back to the data and the tool version behind it is exactly
what this project exists not to produce. Where a format has a metadata channel
the manifest goes in it; where it has none — CSV — it rides in comment lines;
where neither fits, it ships beside the data in the same archive.

Everything is built from the persisted `Run.outcome`, so an export replays no
analysis and asks no model anything. Exporting a run from last week costs one
SELECT.

Chart *images* are deliberately not here. The fourteen renderers are
TypeScript, and the only way a server could produce a faithful PNG would be to
reimplement them — so the frontend exports what it is actually displaying, and
this module owns the data and the prose.
"""

import csv
import io
import json
import zipfile
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from econometrica.db.models import Run

#: Sheet and file names are capped by the format, not by taste: Excel refuses a
#: sheet name over 31 characters and one containing []:*?/\.
_SHEET_LIMIT = 31
_ILLEGAL_SHEET = set(r"[]:*?/\ ")


def export_filename(run: Run, suffix: str) -> str:
    """A name that says which run this was, safe on every filesystem."""
    stamp = run.created_at.astimezone(UTC).strftime("%Y%m%d-%H%M%S")
    return f"econometrica-run-{stamp}-{str(run.id)[:8]}{suffix}"


def manifest_of(run: Run) -> dict[str, Any]:
    """What produced this run, flat enough to read without the whole outcome.

    Separate from the outcome rather than a slice of it: this is the file a
    reader opens first to ask "where did these numbers come from", and burying
    it inside the full payload would defeat that.
    """
    outcome: dict[str, Any] = run.outcome or {}
    quality = outcome.get("quality") or {}
    steps = (outcome.get("execution") or {}).get("outcomes", [])

    return {
        "run_id": str(run.id),
        "question": run.question,
        "status": run.status,
        "tier": run.tier,
        "created_at": run.created_at.isoformat(),
        "exported_at": datetime.now(UTC).isoformat(),
        "data": {
            "source": quality.get("source", ""),
            "tickers": quality.get("tickers", []),
            "start": quality.get("start"),
            "end": quality.get("end"),
            "rows": quality.get("rows"),
            "frequency": quality.get("frequency"),
            "return_method": quality.get("return_method"),
            "fingerprint": quality.get("fingerprint", ""),
            # Carried verbatim: a reader must be able to see that the prices
            # were generated rather than observed.
            "flags": quality.get("flags", []),
        },
        "tools": [
            {
                "step_id": step.get("step_id"),
                "tool": step.get("tool"),
                "status": step.get("status"),
                **_tool_manifest(step.get("result")),
            }
            for step in steps
        ],
    }


def _tool_manifest(result: dict[str, Any] | None) -> dict[str, Any]:
    manifest: dict[str, Any] = (result or {}).get("manifest") or {}
    return {
        "tool_version": manifest.get("tool_version", ""),
        "data_fingerprint": manifest.get("data_fingerprint", ""),
        "params_hash": manifest.get("params_hash", ""),
        "library_versions": manifest.get("library_versions", {}),
        "seed": manifest.get("seed"),
    }


def to_json(run: Run) -> bytes:
    """The whole run: what was asked, what ran, and everything it produced."""
    payload = {
        "run": {
            "id": str(run.id),
            "question": run.question,
            "status": run.status,
            "tier": run.tier,
            "revisions": run.revisions,
            "error": run.error,
            "created_at": run.created_at.isoformat(),
            "input_tokens": run.input_tokens,
            "output_tokens": run.output_tokens,
            "cost_usd": run.cost_usd,
            "latency_ms": run.latency_ms,
        },
        "manifest": manifest_of(run),
        "outcome": run.outcome or {},
    }
    return json.dumps(payload, indent=2, default=str).encode()


def to_markdown(run: Run) -> bytes:
    """A report to paste into a document, with its provenance attached."""
    outcome: dict[str, Any] = run.outcome or {}
    manifest = manifest_of(run)
    data = manifest["data"]
    lines = [f"# {run.question}", ""]

    lines += [
        f"*{run.status}* · {run.tier} tier · {run.created_at.strftime('%Y-%m-%d %H:%M')} UTC",
        "",
    ]

    for flag in data["flags"]:
        if flag.get("severity") in {"risk", "warning"}:
            # First, and unmissable. A report of generated prices that reads
            # like a report of market prices is the failure to avoid.
            code = str(flag.get("code", "")).replace("_", " ")
            lines += [f"> **{code}** — {flag.get('detail', '')}", ""]

    narration = outcome.get("narration") or {}
    narrative = narration.get("narrative") or {}
    lines += ["## Interpretation", ""]
    if narration.get("published") and narrative.get("prose"):
        lines += [str(narrative["prose"]), ""]
    else:
        grounding = narration.get("grounding") or {}
        lines += [
            "No interpretation was published: the draft cited numbers no computed result"
            " supports, so it was withheld whole rather than edited.",
            "",
        ]
        for issue in grounding.get("issues", []):
            lines.append(f"- `{issue.get('text')}` in “{str(issue.get('sentence', '')).strip()}”")
        if grounding.get("issues"):
            lines.append("")

    lines += ["## What ran", "", "| step | tool | status | detail |", "|---|---|---|---|"]
    for step in (outcome.get("execution") or {}).get("outcomes", []):
        refusals = "; ".join(
            verdict.get("detail", "")
            for verdict in step.get("verdicts", [])
            if verdict.get("judged") and not verdict.get("allowed")
        )
        lines.append(
            f"| {step.get('step_id')} | {step.get('tool')} | {step.get('status')} |"
            f" {refusals or step.get('error') or ''} |"
        )
    lines.append("")

    diagnostics = outcome.get("diagnostics") or []
    if diagnostics:
        lines += [
            "## Diagnostics",
            "",
            "| check | statistic | p value | verdict |",
            "|---|---|---|---|",
        ]
        for diagnostic in diagnostics:
            passed = diagnostic.get("passed")
            verdict = "not judged" if passed is None else ("passed" if passed else "failed")
            lines.append(
                f"| {diagnostic.get('name')} | {_round(diagnostic.get('statistic'))} |"
                f" {_round(diagnostic.get('p_value'))} | {verdict} |"
            )
        lines.append("")

    lines += [
        "## Provenance",
        "",
        f"- Data source: `{data['source'] or 'unknown'}`",
        f"- Tickers: {', '.join(data['tickers']) or 'none'}",
        f"- Window: {data['start']} to {data['end']} ({data['rows']} rows, {data['frequency']})",
        f"- Data fingerprint: `{data['fingerprint'] or 'none recorded'}`",
        "",
        "| step | tool | version | data fingerprint | params |",
        "|---|---|---|---|---|",
    ]
    for tool in manifest["tools"]:
        if not tool["data_fingerprint"]:
            # A refused or failed step has no result and therefore no manifest.
            # Blank cells would read as missing provenance rather than as a
            # step that deliberately produced nothing.
            lines.append(
                f"| {tool['step_id']} | {tool['tool']} | — | *{tool['status']}, no result* | — |"
            )
            continue
        lines.append(
            f"| {tool['step_id']} | {tool['tool']} | {tool['tool_version']} |"
            f" `{tool['data_fingerprint'][:16]}` | `{tool['params_hash'][:12]}` |"
        )
    lines += ["", f"Exported {manifest['exported_at']} from run `{run.id}`.", ""]

    return "\n".join(lines).encode()


def _round(value: Any) -> str:
    if not isinstance(value, (int, float)):
        return "—"
    return f"{float(value):.4g}"


def to_csv(run: Run) -> bytes:
    """Every series the run produced, long-format, behind comment lines.

    Long rather than wide because the steps' series do not share an index —
    an ACF is indexed by lag and a volatility path by date — and a wide frame
    would have to invent alignment between them.
    """
    manifest = manifest_of(run)
    buffer = io.StringIO()

    for line in json.dumps(manifest, indent=2, default=str).splitlines():
        buffer.write(f"# {line}\n")

    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(["step_id", "tool", "series", "x", "y"])
    # `or {}` rather than a default: a run that failed before executing stores
    # `execution: null`, so the key is present and `.get(k, {})` hands back None.
    for step in ((run.outcome or {}).get("execution") or {}).get("outcomes", []):
        result = step.get("result") or {}
        for name, series in (result.get("series") or {}).items():
            for x, y in zip(series.get("x", []), series.get("y", []), strict=False):
                writer.writerow([step.get("step_id"), step.get("tool"), name, x, y])

    return buffer.getvalue().encode()


def to_xlsx(run: Run) -> bytes:
    """A workbook: the manifest first, then a sheet per result."""
    book = Workbook()
    book.remove(book.active)

    sheet = book.create_sheet("manifest")
    sheet.append(["field", "value"])
    for key, value in _flatten(manifest_of(run)):
        sheet.append([key, value])
    _widen(sheet)

    used: set[str] = set()
    # `or {}` rather than a default: a run that failed before executing stores
    # `execution: null`, so the key is present and `.get(k, {})` hands back None.
    for step in ((run.outcome or {}).get("execution") or {}).get("outcomes", []):
        result = step.get("result") or {}
        name = _sheet_name(f"{step.get('step_id')}-{step.get('tool')}", used)
        data = book.create_sheet(name)

        estimates = result.get("estimates") or []
        if estimates:
            data.append(["estimate", "value", "std error", "t", "p", "ci low", "ci high"])
            for estimate in estimates:
                data.append(
                    [
                        estimate.get("name"),
                        estimate.get("value"),
                        estimate.get("std_error"),
                        estimate.get("t_stat"),
                        estimate.get("p_value"),
                        estimate.get("ci_low"),
                        estimate.get("ci_high"),
                    ]
                )
            data.append([])

        scalars = result.get("scalars") or {}
        if scalars:
            data.append(["scalar", "value"])
            for key, value in scalars.items():
                data.append([key, value])
            data.append([])

        series = result.get("series") or {}
        if series:
            names = list(series)
            data.append(["series", "x", "y"])
            for name_ in names:
                one = series[name_]
                for x, y in zip(one.get("x", []), one.get("y", []), strict=False):
                    data.append([name_, str(x), y])

        if data.max_row == 1 and data.max_column == 1:
            data.append([f"{step.get('tool')} produced no values ({step.get('status')})"])
        _widen(data)

    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def to_zip(run: Run) -> bytes:
    """Everything, with the manifest as its own file at the root."""
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest_of(run), indent=2, default=str))
        archive.writestr("run.json", to_json(run))
        archive.writestr("report.md", to_markdown(run))
        archive.writestr("series.csv", to_csv(run))
        archive.writestr("results.xlsx", to_xlsx(run))
    return stream.getvalue()


# --- helpers ----------------------------------------------------------------


def _flatten(payload: Any, prefix: str = "") -> list[tuple[str, Any]]:
    """Nested dicts as dotted keys, so a spreadsheet can hold them in two columns."""
    rows: list[tuple[str, Any]] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            rows.extend(_flatten(value, f"{prefix}.{key}" if prefix else str(key)))
    elif isinstance(payload, list):
        for index, value in enumerate(payload):
            rows.extend(_flatten(value, f"{prefix}[{index}]"))
    else:
        rows.append((prefix, payload))
    return rows


def _sheet_name(candidate: str, used: set[str]) -> str:
    cleaned = "".join("-" if char in _ILLEGAL_SHEET else char for char in candidate)
    name = cleaned[:_SHEET_LIMIT] or "sheet"
    suffix = 2
    while name in used:
        tail = f"-{suffix}"
        name = cleaned[: _SHEET_LIMIT - len(tail)] + tail
        suffix += 1
    used.add(name)
    return name


def _widen(sheet: Any) -> None:
    """Column widths from the content, so nothing opens as ####."""
    for column in range(1, sheet.max_column + 1):
        longest = max(
            (
                len(str(sheet.cell(row=row, column=column).value or ""))
                for row in range(1, sheet.max_row + 1)
            ),
            default=10,
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(60, max(10, longest + 2))
